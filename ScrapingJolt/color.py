import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# 入出力ファイル
INPUT_EXCEL = "jolnet_transfer_data_sorted.xlsx"
OUTPUT_EXCEL = "jolnet_transfer_data_colored.xlsx"

# 地域ごとの背景色（Excelのカラーコード）
region_colors = {
    "北海道地方": "FFEBEE",
    "東北地方": "FFFDE7",
    "関東地方": "E3F2FD",
    "中部地方": "E8F5E9",
    "近畿地方": "F3E5F5",
    "中国地方": "FBE9E7",
    "四国地方": "E0F2F1",
    "九州地方": "FFF3E0",
    "沖縄地方": "E1F5FE",
    "海外地方": "ECEFF1",
    "不明":     "FFFFFF"
}

# ヘッダーの色（薄グレー）
header_fill = PatternFill(start_color="CFD8DC", end_color="CFD8DC", fill_type="solid")

# 削除するカラム名
columns_to_remove = ["学校区分", "詳細", "入試科目等", "出願期間等"]

# Excel読み込み
if not os.path.exists(INPUT_EXCEL):
    raise FileNotFoundError(f"{INPUT_EXCEL} が見つかりません。")

wb = load_workbook(INPUT_EXCEL)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    header_row = 1
    max_col = ws.max_column
    max_row = ws.max_row

    # --- ヘッダー読み込み ---
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, max_col + 1)]

    # --- 削除対象列のインデックス取得 ---
    cols_to_delete = []
    for col_idx, header in enumerate(headers, start=1):
        if header in columns_to_remove:
            cols_to_delete.append(col_idx)

    # --- 後ろから順に削除（列番号がずれるため） ---
    for col_idx in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(col_idx)

    # --- ヘッダー再取得（列削除後） ---
    max_col = ws.max_column
    headers = [ws.cell(row=header_row, column=col).value for col in range(1, max_col + 1)]

    # --- 地域列を探す ---
    region_col = None
    for col in range(1, max_col + 1):
        if ws.cell(row=header_row, column=col).value == "地域":
            region_col = col
            break

    if region_col is None:
        raise ValueError(f"{sheet_name} シートで「地域」列が見つかりません。")

    # ✅ ヘッダーに色を付け、固定
    for col in range(1, max_col + 1):
        ws.cell(row=header_row, column=col).fill = header_fill

    ws.freeze_panes = "A2"  # ヘッダー固定

    # ✅ 各行に地域に応じた色を適用
    for row in range(header_row + 1, ws.max_row + 1):
        region = ws.cell(row=row, column=region_col).value
        color = region_colors.get(region, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).fill = fill

# 保存
wb.save(OUTPUT_EXCEL)
print(f"✅ 書式付きExcelを保存しました：{OUTPUT_EXCEL}")